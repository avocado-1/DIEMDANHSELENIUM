import openpyxl


class LoginData:

    @staticmethod
    def getLoginData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\PENGUIN\\Downloads\\LoginData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=i).value == test_case_name:
                for j in range(2, sheet.max_row + 1):
                    Dict[sheet.cell(row=j, column=1).value] = sheet.cell(row=j, column=i).value
        return [Dict]

